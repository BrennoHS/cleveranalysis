"""
report_file_parser.py - Deterministic parsing for structured publisher report files.
Supports .xlsx and .csv, selecting metrics by an analysis date range.
"""

from __future__ import annotations

import csv
import io
import re
from datetime import date, datetime
from typing import Any

from openpyxl import load_workbook


DATE_HEADERS = {
    "date",
    "day",
    "report date",
}

SERVED_HEADERS = {
    "impressions",
    "served impressions",
    "ad server impressions",
    "total impressions",
    "measurable impressions",
}

VIEWABLE_HEADERS = {
    "viewable impressions",
    "active view viewable impressions",
    "active views",
    "active view",
}


def _normalize_header(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = re.sub(r"\s+", " ", text)
    text = re.sub(r"[^a-z0-9 ]", "", text)
    return text.strip()


def _normalize_number(value: Any) -> int | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return int(value)

    text = str(value).strip()
    if not text:
        return None

    digits = re.sub(r"[^\d]", "", text)
    if not digits:
        return None
    return int(digits)


def _parse_date_value(value: Any) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    text = str(value).strip()
    if not text:
        return None

    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%Y/%m/%d", "%d-%m-%Y", "%m-%d-%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue

    return None


def _find_header_row(rows: list[list[Any]]) -> tuple[int, list[str]]:
    for idx, row in enumerate(rows):
        normalized = [_normalize_header(cell) for cell in row]
        if not any(normalized):
            continue

        has_date = any(cell in DATE_HEADERS for cell in normalized)
        has_served = any(cell in SERVED_HEADERS for cell in normalized)
        has_viewable = any(cell in VIEWABLE_HEADERS for cell in normalized)
        if has_date and (has_served or has_viewable):
            return idx, normalized

    raise ValueError(
        "Could not identify table headers. Expected columns like Date, Impressions and Viewable Impressions."
    )


def _index_by_header(headers: list[str], candidates: set[str]) -> int | None:
    for idx, header in enumerate(headers):
        if header in candidates:
            return idx
    return None


def _rows_from_xlsx(content: bytes) -> list[list[Any]]:
    workbook = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    sheet = workbook.active
    rows: list[list[Any]] = []
    for row in sheet.iter_rows(values_only=True):
        rows.append(list(row))
    return rows


def _rows_from_csv(content: bytes) -> list[list[Any]]:
    text = content.decode("utf-8-sig", errors="ignore")
    reader = csv.reader(io.StringIO(text))
    return [list(row) for row in reader]


def _extract_metrics_for_period(rows: list[list[Any]], start_date: date, end_date: date) -> dict:
    header_idx, headers = _find_header_row(rows)

    date_col = _index_by_header(headers, DATE_HEADERS)
    served_col = _index_by_header(headers, SERVED_HEADERS)
    viewable_col = _index_by_header(headers, VIEWABLE_HEADERS)

    if date_col is None:
        raise ValueError("Date column not found in uploaded report.")
    if served_col is None:
        raise ValueError("Impressions column not found in uploaded report.")

    served_total = 0
    viewable_total = 0
    match_count = 0

    for row in rows[header_idx + 1 :]:
        if date_col >= len(row):
            continue

        row_date = _parse_date_value(row[date_col])
        if row_date is None or row_date < start_date or row_date >= end_date:
            continue

        served_val = _normalize_number(row[served_col]) if served_col < len(row) else None
        viewable_val = _normalize_number(row[viewable_col]) if viewable_col is not None and viewable_col < len(row) else 0

        if served_val is None:
            continue

        served_total += served_val
        viewable_total += viewable_val or 0
        match_count += 1

    if match_count == 0:
        raise ValueError(
            f"No rows found for analysis period [{start_date.isoformat()}, {end_date.isoformat()}) in uploaded report."
        )

    return {
        "served_impressions": served_total,
        "viewable_impressions": viewable_total,
        "start_date": start_date.isoformat(),
        "end_date": end_date.isoformat(),
    }


def parse_publisher_report_file(content: bytes, filename: str, start_date_str: str, end_date_str: str) -> dict:
    start_date = _parse_date_value(start_date_str)
    end_date = _parse_date_value(end_date_str)
    if start_date is None:
        raise ValueError("start_date must be a valid date in YYYY-MM-DD format.")
    if end_date is None:
        raise ValueError("end_date must be a valid date in YYYY-MM-DD format.")
    if start_date >= end_date:
        raise ValueError("start_date must be earlier than end_date.")

    name = (filename or "").lower()
    if name.endswith(".xlsx"):
        rows = _rows_from_xlsx(content)
    elif name.endswith(".csv"):
        rows = _rows_from_csv(content)
    else:
        raise ValueError("Unsupported file type. Please upload .xlsx or .csv report files.")

    return _extract_metrics_for_period(rows, start_date, end_date)
